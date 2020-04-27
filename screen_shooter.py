import webbrowser
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import tkinter
from tkinter import Label, Button, Entry, Checkbutton, OptionMenu, Canvas, Frame
from tkinter import StringVar, IntVar
from tkinter import Listbox
from tkinter import Scrollbar
from tkinter import messagebox
from tkinter import N, S, E, W
from tkinter import VERTICAL
from tkinter import Toplevel
from tkinter.filedialog import askopenfilename
from threading import Thread
from tkinter.messagebox import showinfo

'''Below module allows us to interact with Windows files.'''
import os

'''below 3 lines allows script to check the directory where it is executed, so it knows where to crete the excel file. I copied the whole block from stack overflow'''
abspath = os.path.abspath(__file__)
current_directory = os.path.dirname(abspath)
os.chdir(current_directory)

'''Below module allows us to interact with Excel files'''
import openpyxl
from openpyxl import Workbook, load_workbook

SCREENSHOT_FOLDER_NAME = ""

stop_test = 0

row_counter = 0

test_sequence_source_excel = ""

dictionary_of_action_index_per_row = {}
dictionary_of_action_selector_per_row = {}
dictionary_of_action_selector_variable_per_row = {}
dictionary_of_action_input_per_row = {}


list_of_actions = ["Open (URL)", "Click (CSS selector)", "Send (text)", "Set resolution (width)", "Screenshot (save as)", "Wait (seconds)", "Open (URL) +", "Click (CSS selector) +", "Send (text) +", "Screenshot (save as) +", "Set resolution (width) +"]
list_of_browsers = ["PC browser 1920x1080", "Mobile browser 768x1204"]

def get_files_in_script_directory():
    '''Get file names in directory'''
    file_names = []
    for root, dirs, files in os.walk(current_directory):
        for filename in files:
            file_names.append(filename)
    return file_names

def check_if_need_to_create_folder(dictionary_of_sequences):
    need_to_create_folder = 0
    for sequence in dictionary_of_sequences:
        for action, value in dictionary_of_sequences[sequence].items():
            if dictionary_of_sequences[sequence][action]["action"] == "Screenshot (save as)" or dictionary_of_sequences[sequence][action]["action"] == "Screenshot (save as) +":
                need_to_create_folder = 1
    return need_to_create_folder

def create_single_folder():
    global SCREENSHOT_FOLDER_NAME
    provided_folder_name = entry_folder_name.get()
    if provided_folder_name == "" or folder_name_option_var.get() == "Folder name +":
        try:
            time_stamp_minutes = str(time.strftime("%m-%d %Hh%M"))
            SCREENSHOT_FOLDER_NAME = "./Screenshots " + time_stamp_minutes + "/"
            os.makedirs(SCREENSHOT_FOLDER_NAME)
        except:
            time_stamp_seconds = str(time.strftime("%Hh%Ms%S"))
            SCREENSHOT_FOLDER_NAME = "./Screenshots " + time_stamp_seconds + "/"
            os.makedirs(SCREENSHOT_FOLDER_NAME)
    else:
        try:
            SCREENSHOT_FOLDER_NAME = "./" + provided_folder_name + "/"
            os.makedirs(SCREENSHOT_FOLDER_NAME)
        except:
            time_stamp_minutes = str(time.strftime("%m-%d %Hh%M"))
            SCREENSHOT_FOLDER_NAME = "./Screenshots " + time_stamp_minutes + "/"
            os.makedirs(SCREENSHOT_FOLDER_NAME)


def create_folders_for_screenshots_if_any_screenshots_are_taken(dictionary_of_sequences):
    global SCREENSHOT_FOLDER_NAME
    need_to_create_folder = check_if_need_to_create_folder(dictionary_of_sequences)
    if need_to_create_folder == 1:
        create_single_folder()
        if folder_name_option_var.get() == "Folder name +":
            list_of_subfolder_names = []
            wb = load_workbook(test_sequence_source_excel)
            ws = wb.worksheets[1]
            needed_column = 0
            current_column = 1
            columns_to_check = ws.cell(row=1, column=current_column)
            while columns_to_check.value != None:
                if columns_to_check.value == entry_folder_name.get():
                    needed_column = current_column
                current_column += 1
                columns_to_check = ws.cell(row=1, column=current_column)
            needed_row = 2
            needed_value_cell = ws.cell(row=needed_row, column=needed_column)
            while needed_value_cell.value != None:
                if needed_value_cell.value not in list_of_subfolder_names:
                    list_of_subfolder_names.append(needed_value_cell.value)
                needed_row += 1
                needed_value_cell = ws.cell(row=needed_row, column=needed_column)
            for sub_folder_name in list_of_subfolder_names:
                try:
                    os.makedirs(SCREENSHOT_FOLDER_NAME + sub_folder_name + "/")
                except:
                    insert_text("Was unable to create subfolder " + SCREENSHOT_FOLDER_NAME + sub_folder_name + "/")

def look_up_action_value_from_excel(single_value, test_reiteration):
    wb = load_workbook(test_sequence_source_excel)
    ws = wb.worksheets[1]
    needed_column = 0
    needed_row = test_reiteration + 2
    current_column = 1
    columns_to_check = ws.cell(row=1, column=current_column)
    while columns_to_check.value != None:
        if columns_to_check.value == single_value:
            needed_column = current_column
        current_column = current_column + 1
        columns_to_check = ws.cell(row=1, column=current_column)
    needed_value_cell = ws.cell(row=needed_row, column=needed_column)
    # while needed_value_cell.value == None: # this part would force to skip empty values
    #    needed_row -= 1
    #    needed_value_cell = ws.cell(row=needed_row, column=needed_column)
    return needed_value_cell.value


def build_single_test_sequence(test_reiteration):
    list_of_complex_inputs = ["Open (URL) +", "Click (CSS selector) +", "Send (text) +", "Screenshot (save as) +", "Language +", "Browser +"]
    dictionary_of_single_test_sequence = {}
    for i in range(row_counter):
        dictionary_of_single_test_sequence[i] = {}
        single_action = dictionary_of_action_selector_variable_per_row[i].get()
        dictionary_of_single_test_sequence[i]["action"] = single_action
        single_value = dictionary_of_action_input_per_row[i].get()
        if single_action in list_of_complex_inputs:
            complex_value = look_up_action_value_from_excel(single_value, test_reiteration)
            dictionary_of_single_test_sequence[i]["value"] = complex_value
        else:
            dictionary_of_single_test_sequence[i]["value"] = single_value
    return dictionary_of_single_test_sequence

def check_if_excel_file_is_selected_and_create_one_if_not():
    if test_sequence_source_excel == "":
        save_testing_sequence_popup()

def get_the_number_of_test_sequences():
    number_of_test_sequences = 0
    try:
        wb = load_workbook(test_sequence_source_excel)
        ws = wb.worksheets[1]
        number_of_columns = 1
        columns_to_check = ws.cell(row=1, column=number_of_columns)
        while columns_to_check.value != None:
            number_of_columns = number_of_columns + 1
            columns_to_check = ws.cell(row=1, column=number_of_columns)
        for column_to_check in range(1, number_of_columns):
            number_of_rows = 1
            rows_to_check = ws.cell(row=number_of_rows, column=column_to_check)
            while rows_to_check.value != None:
                number_of_rows = number_of_rows + 1
                rows_to_check = ws.cell(row=number_of_rows, column=column_to_check)
            if number_of_test_sequences < number_of_rows:
                number_of_test_sequences = number_of_rows
        number_of_test_sequences -= 2
    except:
        number_of_test_sequences = 1
    return number_of_test_sequences



def remove_action_row():
    global row_counter
    if row_counter > 0:
        row_counter -= 1
        dictionary_of_action_index_per_row[row_counter].destroy()
        dictionary_of_action_selector_per_row[row_counter].destroy()
        dictionary_of_action_input_per_row[row_counter].destroy()
        del dictionary_of_action_selector_variable_per_row[row_counter]

def add_action_row():
    global row_counter
    global list_of_actions
    action_row = (row_counter) % 20
    action_column = row_counter // 20  * 3 
    dictionary_of_action_index_per_row[row_counter] = Label(main_window_of_gui, text = str(row_counter + 1))
    dictionary_of_action_index_per_row[row_counter].grid(row = action_row + 3, column = action_column)
    dictionary_of_action_selector_variable_per_row[row_counter] = StringVar()
    dictionary_of_action_selector_variable_per_row[row_counter].set(list_of_actions[0])
    dictionary_of_action_selector_per_row[row_counter] = OptionMenu(main_window_of_gui, dictionary_of_action_selector_variable_per_row[row_counter], *list_of_actions)
    dictionary_of_action_selector_per_row[row_counter].config(width=18)
    dictionary_of_action_selector_per_row[row_counter].grid(row = action_row + 3, column = action_column + 1, columnspan = 1)
    dictionary_of_action_input_per_row[row_counter] = Entry(main_window_of_gui)
    dictionary_of_action_input_per_row[row_counter].config(width=55)
    dictionary_of_action_input_per_row[row_counter].grid(row = action_row + 3, column = action_column + 2, columnspan = 1)
    row_counter += 1


def go_to(driver, value):
    insert_text("Opening " + value)
    driver.get(value)

def click_element(driver, value):
    insert_text("Clicking " + value)
    driver.find_element_by_css_selector(value).click()

def get_last_clicked_selector(dictionary_of_current_test_itiration, action_index):
    last_clicked_selector = ""
    for i in range(action_index):
        if dictionary_of_current_test_itiration[i]["action"] == "Click (CSS selector)" or dictionary_of_current_test_itiration[i]["action"] == "Click (CSS selector) +":
            last_clicked_selector = dictionary_of_current_test_itiration[i]["value"]
    return last_clicked_selector

def enter_text(driver, value, dictionary_of_current_test_itiration, action_index):
    last_clicked_selector = get_last_clicked_selector(dictionary_of_current_test_itiration, action_index)
    input_text = value
    insert_text("Typing " + '"' + input_text + '"' + " in element " + last_clicked_selector)
    driver.find_element_by_css_selector(last_clicked_selector).send_keys(input_text)

def create_browser(test_reiteration):
    pc_browser_width = 1920
    pc_browser_height = 6000
    mobile_browser_width = 768
    mobile_browser_height = 6000
    options = webdriver.ChromeOptions()
    if display_browser_var.get() == 0:
        options.add_argument('headless')
        insert_text('Opening "hidden" browser.')
    if browser_type_var.get() == "PC browser 1920x1080":
        selected_width = pc_browser_width
        selected_height = pc_browser_height
    if browser_type_var.get() == "Mobile browser 768x1204":
        selected_width = mobile_browser_width
        selected_height = mobile_browser_height
    driver = webdriver.Chrome(options=options)
    driver.set_window_size(selected_width, selected_height)
    driver.implicitly_wait(1)
    return driver

def set_resolution(driver, value):
    web_browser_window_width = int(value)
    web_browser_window_height = 10000
    driver.set_window_size(web_browser_window_width, web_browser_window_height)

def take_screenshot(driver, value, action_index, test_reiteration):
    global SCREENSHOT_FOLDER_NAME
    if folder_name_option_var.get() == "Folder name +":
        sub_folder_name_look_up_key = entry_folder_name.get()
        sub_folder_name = look_up_action_value_from_excel(sub_folder_name_look_up_key, test_reiteration)
        save_to = SCREENSHOT_FOLDER_NAME + sub_folder_name + "/"
    else:
        save_to = SCREENSHOT_FOLDER_NAME
    screenshot_name = value
    if ".png" not in screenshot_name:
        screenshot_name += ".png"
    if dictionary_of_action_selector_variable_per_row[action_index].get() != "Screenshot (save as) +":
        screenshot_name = str(test_reiteration + 1) + "-" + str(action_index + 1) + " " + screenshot_name
    insert_text("Taking screenshot: " + screenshot_name)
    screenshot_name = save_to + screenshot_name
    whole_page = driver.find_element_by_tag_name('body')
    whole_page.screenshot(screenshot_name)

def single_action(driver, action_value_tuple, dictionary_of_current_test_itiration, action_index, test_reiteration):
    action = action_value_tuple["action"]
    value = str(action_value_tuple["value"])
    if action == "Open (URL)" or action == "Open (URL) +":
        try:
            go_to(driver, value)
        except:
            insert_text("Run number " + str(test_reiteration) + ". Was unable to open " + value)
    if action == "Click (CSS selector)" or action == "Click (CSS selector) +":
        try:
            click_element(driver, value)
        except:
            insert_text("Run number " + str(test_reiteration) + ". Was unable to click " + value)
    if action == "Send (text)" or action == "Send (text) +":
        try:
            enter_text(driver, value, dictionary_of_current_test_itiration, action_index)
        except:
            insert_text("Run number " + str(test_reiteration) + ". Was unable to write " + value)
    if action == "Set resolution (width)":
        try:
            set_resolution(driver, value)
        except:
            insert_text("Run number " + str(test_reiteration) + ". Was unable to set resolution width to " + value)
    if action == "Screenshot (save as)" or action == "Screenshot (save as) +":
        try:
            take_screenshot(driver, value, action_index, test_reiteration)
        except:
            insert_text("Run number " + str(test_reiteration) + ". Was unable to take screenshot " + value)
    if "Wait" in action:
        try:
            time.sleep(int(value))
        except:
            insert_text("Run number " + str(test_reiteration) + ". Was unable to wait " + str(value) + " second(s).")

def perform_actions(dictionary_of_current_test_itiration, test_reiteration):
    global stop_test
    driver = create_browser(test_reiteration)
    for action_index, action_value_tuple in dictionary_of_current_test_itiration.items():
        if stop_test == 0:
            single_action(driver, action_value_tuple, dictionary_of_current_test_itiration, action_index, test_reiteration)
        else:
            insert_text("/!\\ Sequence aborted /!\\")
    driver.quit()


def insert_text(text):
    text_box.insert('end', text)
    text_box.see("end")

def open_apps_folder():
    os.startfile(current_directory + "\\", 'open') 

def save_testing_sequence(entry_test_sequence_name, popup_provide_test_sequence_name):
    global test_sequence_source_excel
    global row_counter
    test_sequence_name = entry_test_sequence_name.get()
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.cell(row=1, column=1, value=folder_name_option_var.get())
    ws.cell(row=1, column=2, value=entry_folder_name.get())
    for line_number in range(row_counter):
        ws.cell(row=(line_number + 2), column=1, value=dictionary_of_action_selector_variable_per_row[line_number].get())
        ws.cell(row=(line_number + 2), column=2, value=dictionary_of_action_input_per_row[line_number].get())
    if ".xlsx" not in test_sequence_name:
        test_sequence_name += ".xlsx"
    wb.save(test_sequence_name)
    test_sequence_source_excel = test_sequence_name
    popup_provide_test_sequence_name.destroy()

def get_main_window_of_gui_postion():
    x_coordinate = main_window_of_gui.winfo_x()
    y_coordinate = main_window_of_gui.winfo_y()
    return ("230x80" + "+" + (str(x_coordinate) + "+" + str(y_coordinate)))

def save_testing_sequence_popup():
    popup_provide_test_sequence_name = Toplevel()
    popup_provide_test_sequence_name.geometry(get_main_window_of_gui_postion())
    popup_provide_test_sequence_name.wm_attributes("-topmost", 1)
    popup_provide_test_sequence_name.wm_title("Test sequence name")
    label_test_sequence_name = Label(popup_provide_test_sequence_name, text = "Please, provide the name of this test run:")
    label_test_sequence_name.pack()
    entry_test_sequence_name = Entry(popup_provide_test_sequence_name, width=30)
    entry_test_sequence_name.pack()
    entry_test_sequence_name.focus()
    button_save_test_sequence_name = Button(popup_provide_test_sequence_name, text="Save test sequence", width = 30, height = 3, command = lambda: save_testing_sequence(entry_test_sequence_name, popup_provide_test_sequence_name))
    button_save_test_sequence_name.pack()
    popup_provide_test_sequence_name.mainloop()

def find_the_row_of_the_next_empty_cell():
    number_of_lines = 1
    wb = load_workbook(test_sequence_source_excel)
    ws = wb.worksheets[0]
    cell_to_check = ws.cell(row = number_of_lines, column = 1)
    while cell_to_check.value != None:
        number_of_lines = number_of_lines + 1
        cell_to_check = ws.cell(row = number_of_lines, column = 1)
    return number_of_lines, ws

def build_test_sequence_with_import():
    global row_counter
    while row_counter > 0:
        remove_action_row()
    number_of_lines, ws = find_the_row_of_the_next_empty_cell()
    folder_name_option_var.set(ws.cell(row=1, column=1).value)
    entry_folder_name.delete(0,'end')
    try:
        entry_folder_name.insert(0, ws.cell(row=1, column=2).value)
    except:
        entry_folder_name.insert(0, "")
    for line in range(number_of_lines - 2):
        add_action_row()
        dictionary_of_action_selector_variable_per_row[line].set(ws.cell(row=line+2, column=1).value)
        dictionary_of_action_input_per_row[line].insert(0, ws.cell(row=line+2, column=2).value)
    

def choose_test_sequence_source_excel():
    global test_sequence_source_excel
    test_sequence_source_excel = askopenfilename(filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    build_test_sequence_with_import()

def stop_the_test():
    insert_text("Aborting current run!")
    global stop_test
    stop_test = 1

def run_selected_actions():
    def slow_magic():
        global stop_test
        list_of_sequences = []
        stop_test = 0
        text_box.delete(0,'end')
        insert_text("Starting the process!")
        insert_text("Please, don't interact with interface until done. You can minimize this window.")
        dictionary_of_sequences = {}
        number_of_test_sequences = get_the_number_of_test_sequences()
        for test_reiteration in range(number_of_test_sequences):
            dictionary_of_sequences[test_reiteration] = build_single_test_sequence(test_reiteration)
        create_folders_for_screenshots_if_any_screenshots_are_taken(dictionary_of_sequences)
        for test_reiteration in range(number_of_test_sequences):
            t = Thread(target = perform_actions, args = (dictionary_of_sequences[test_reiteration], test_reiteration))
            list_of_sequences.append(t)
            t.start()
        for t in list_of_sequences:t.join()
        insert_text("****************** ALL ACTIONS PERFORMED ******************")
    executing = Thread(target=slow_magic)
    executing.start()

def _onKeyRelease(event):
    ctrl  = (event.state & 0x4) != 0
    if event.keycode==88 and  ctrl and event.keysym.lower() != "x": 
        event.widget.event_generate("<<Cut>>")

    if event.keycode==86 and  ctrl and event.keysym.lower() != "v": 
        event.widget.event_generate("<<Paste>>")

    if event.keycode==67 and  ctrl and event.keysym.lower() != "c":
        event.widget.event_generate("<<Copy>>")

    if event.keycode==65 and  ctrl and event.keysym.lower() != "a":
        event.widget.event_generate("<<SelectAll>>")

main_window_of_gui = tkinter.Tk()
main_window_of_gui.title("Screen-shooter v18/03/2020")
main_window_of_gui.wm_attributes("-topmost", 1)


main_window_of_gui.bind_all("<Key>", _onKeyRelease, "+")

folder_name_option_var = StringVar()
folder_name_option_list = ["Folder name", "Folder name +"]

browser_type_var = StringVar()
display_browser_var = IntVar()

display_browser_toggle = Checkbutton(main_window_of_gui, text="Display browser", variable=display_browser_var)
display_browser_toggle.grid(row = 0, column = 0, columnspan = 2)

label_browser = Label(main_window_of_gui, text = "Select browser type: ")
label_browser.grid(row = 1, column = 0, columnspan = 2)

browser_type_var.set(list_of_browsers[0])
drop_down_browser_type = OptionMenu(main_window_of_gui, browser_type_var, *list_of_browsers)
drop_down_browser_type.config(width = 22)
drop_down_browser_type.grid(row = 1, column = 2, columnspan = 1)


folder_name_option_var.set(folder_name_option_list[0])
drop_down_folder_name = OptionMenu(main_window_of_gui, folder_name_option_var, *folder_name_option_list)
drop_down_folder_name.config(width=18)
drop_down_folder_name.grid(row = 2, column = 1, columnspan = 1)

entry_folder_name = Entry(main_window_of_gui)
entry_folder_name.config(width=55)
entry_folder_name.grid(row = 2, column = 2, columnspan = 1)

button_add_action_row = Button(main_window_of_gui, text = "Add row", width = 16, height = 3, command = add_action_row)
button_add_action_row.grid(row = 0, column = 2, columnspan = 2)
button_remove_action_row = Button(main_window_of_gui, text = "Remove row", width = 16, height = 3, command = remove_action_row)
button_remove_action_row.grid(row = 0, column = 4, columnspan = 1)
button_perform_actions = Button(main_window_of_gui, text = "Perform actions", width = 16, height = 3, command = run_selected_actions) #perform_actions
button_perform_actions.grid(row = 0, column = 5, rowspan = 1, columnspan = 1)

text_box = Listbox(main_window_of_gui, height=8)
text_box.grid(column=0, row=100, columnspan=13, sticky=(N,W,E,S))  # columnspan − How many columns widgetoccupies; default 1.
main_window_of_gui.grid_columnconfigure(0, weight=1)
main_window_of_gui.grid_rowconfigure(13, weight=1)
#scroll bar
my_scrollbar = Scrollbar(main_window_of_gui, orient=VERTICAL, command=text_box.yview)
my_scrollbar.grid(column=13, row=100, sticky=(N,S))
#attaching scroll bar to text box
text_box['yscrollcommand'] = my_scrollbar.set

button_open_the_folder = Button(main_window_of_gui, text = "Open the folder", width = 16, height = 3, command = open_apps_folder)
button_open_the_folder.grid(row = 0, column = 6, rowspan = 1, columnspan = 1)

button_save_testing_sequence = Button(main_window_of_gui, text = "Save tests", width = 16, height = 3, command = save_testing_sequence_popup)
button_save_testing_sequence.grid(row = 0, column = 7, rowspan = 1, columnspan = 1)

button_load_testing_sequence = Button(main_window_of_gui, text = "Load tests", width = 16, height = 3, command = choose_test_sequence_source_excel)
button_load_testing_sequence.grid(row = 0, column = 8, rowspan = 1, columnspan = 1)

button_stop_the_test = Button(main_window_of_gui, text = "Cancel run", width = 16, height = 3, command = stop_the_test)
button_stop_the_test.grid(row = 0, column = 9, rowspan = 1, columnspan = 1)

main_window_of_gui.mainloop()